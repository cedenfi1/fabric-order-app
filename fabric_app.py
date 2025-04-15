import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import shutil
from datetime import datetime, timedelta

st.set_page_config(page_title="Fabric Order Processor", layout="wide")
st.title("🧵 Fabric Order Processor")

uploaded_file = st.file_uploader("Upload your CSV file", type="csv")

if uploaded_file:
    data = pd.read_csv(uploaded_file, low_memory=False)
    key_columns = ['Order #', 'Customer Name', 'Sku', 'Brand', 'Product Name', 'Color', 'Quantity']
    data = data[[col for col in key_columns if col in data.columns]]

    data['Order #'] = pd.to_numeric(data['Order #'], errors='coerce')
    min_order = int(data['Order #'].min())
    max_order = int(data['Order #'].max())
    order_range_text = f"{min_order} to {max_order}"

    data['Brand'] = data['Brand'].astype(str).str.upper().str.strip()
    data = data[data['Brand'].isin(['FABRIC', 'KIT'])]

    kits = data[data['Brand'] == 'KIT']
    fabrics = data[data['Brand'] == 'FABRIC']

    fabrics_grouped = fabrics.groupby(['Customer Name', 'Sku', 'Brand', 'Product Name'], as_index=False)['Quantity'].sum()
    fabrics_grouped['Color'] = ""

    kits_grouped = kits.groupby(['Customer Name', 'Sku', 'Brand', 'Product Name', 'Color'], as_index=False)['Quantity'].sum()

    main_data = pd.concat([fabrics_grouped, kits_grouped], ignore_index=True)
    main_data = main_data.sort_values(by=['Sku', 'Quantity'])

    main_tally = main_data.groupby(['Sku', 'Brand', 'Product Name', 'Color', 'Quantity']).size().reset_index(name='Count')

    pivot = main_tally.pivot_table(index=['Sku', 'Brand', 'Product Name', 'Color'],
                                   columns='Quantity',
                                   values='Count',
                                   fill_value=0).reset_index()

    new_cols = []
    for col in pivot.columns:
        if isinstance(col, (int, float)):
            yards = 0.5 * col
            label = f"{int(col)} QTY ({yards} yd)"
            new_cols.append(label)
        else:
            new_cols.append(col)
    pivot.columns = new_cols

    qty_cols = [col for col in pivot.columns if "QTY" in col]
    total_quantity = [
        sum(int(col.split()[0]) * row[col] for col in qty_cols)
        for _, row in pivot.iterrows()
    ]

    final_df = pivot[['Sku', 'Product Name', 'Color'] + qty_cols]

    template_path = "Cut Sheet Template (1).xlsx"
    output_path = "cut_sheet_output.xlsx"
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    ws.delete_rows(1)

    sale_date = (datetime.now() - timedelta(days=2)).strftime("%m/%d/%Y")
    ws.merge_cells("H2:I2")
    ws["H2"].value = f"Date of Sale\n{sale_date}"
    ws["H2"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    ws.merge_cells("L2:M2")
    ws["L2"].value = f"Order Range: {order_range_text}"
    ws["L2"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    start_row = 4
    start_col = 1

    for i, row in final_df.iterrows():
        row_index = start_row + i
        for j, value in enumerate(row, start=start_col):
            if isinstance(value, (int, float)) and value == 0 and "QTY" in final_df.columns[j - 1]:
                continue
            cell = ws.cell(row=row_index, column=j)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")
        # Write Total Quantity in the last available column
        ws.cell(row=row_index, column=start_col + len(final_df.columns), value=total_quantity[i])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.success("✅ File processed successfully!")
    st.download_button(
        label="📥 Download Filled Cut Sheet",
        data=output,
        file_name="cut_sheet_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
